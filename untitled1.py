# -*- coding: utf-8 -*-
"""
Created on Fri Jul 17 11:47:38 2020

@author: exzob
"""


from openpyxl import load_workbook
 
workbook = load_workbook(r'C:\Users\exzob\Desktop\flaskversion1\test.xlsx')
#booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
sheets = workbook.get_sheet_names() 
print(sheets)        #从名称获取sheet
booksheet = workbook.get_sheet_by_name(sheets[1])
 
rows = booksheet.rows
columns = booksheet.columns
#迭代所有的行
for row in rows:
    line = [col.value for col in row]
 
#通过坐标读取值

cell_11 = booksheet.cell(row=2, column=1).value
print(cell_11)